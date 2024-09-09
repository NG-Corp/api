#Imports
import pytesseract
from PIL import Image
import re
import pandas as pd
from collections import Counter
from math import sqrt
from difflib import get_close_matches
import cv2
import numpy as np
from PIL import Image
import openpyxl
from flask import Flask, request
#Util Funcs
#Finding
def find(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    d = [sheet.cell(row=i,column=2).value for i in range(1,sheet.max_row+1)]
    n = [sheet.cell(row=i,column=1).value for i in range(1,sheet.max_row+1)]
    t = [sheet.cell(row=i,column=3).value for i in range(1,sheet.max_row+1)]
    find.r = [sheet.cell(row=i,column=3).value for i in range(1,sheet.max_row+1)]
    for i in range(0,int(len(d))):
                find.r[i] = f'{n[i]} {d[i]} {t[i]} /m'
    return find.r
#Scanning
def scan(imng,saveto):
    """ Scans Receipts and finds the total and date. Heart of the ScansApp app.

        ### Documentery ###

        img- the image you want to scan.

        saveto - where you want to save it to (an excel file)

        By Nevaan Gupta at NGCorp
    """
# Image Prep
    imgd = Image.open(imng)
    assert imgd is not None, "file could not be read, check with os.path.exists()"
    #contrast = 1 
    #brightness = 5
    #out = cv2.addWeighted(imgd,contrast,imgd,0,brightness)
    #im_gray = cv2.cvtColor(out,cv2.COLOR_BGR2GRAY)
    #th, imf= cv2.threshold(im_gray, 128, 192,cv2.THRESH_OTSU)  
    #cv2.imwrite("n.jpg",imf)
    #remshad("n.jpg")

    #Mini Functions
    def remshad(img):
        img = cv2.imread(img)
        rgb_planes = cv2.split(img)
        result_norm_planes = []
        for plane in rgb_planes:
            dilated_img = cv2.dilate(plane, np.ones((7,7), np.uint8))
            bg_img = cv2.medianBlur(dilated_img, 21)
            diff_img = 255 - cv2.absdiff(plane, bg_img)
            norm_img = cv2.normalize(diff_img,None, alpha=0, beta=255, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_8UC1)
            result_norm_planes.append(norm_img)
        result_norm = cv2.merge(result_norm_planes)
        cv2.imwrite('p.png', result_norm)
        scan("p.png")
    def find_amounts(text):
        amounts = re.findall(r'\d+\.\d{2}\b', text)
        floats = [float(amount) for amount in amounts]
        unique = list(dict.fromkeys(floats))
        return unique
    def word2vec(word):
        # count the characters in word
        cw = Counter(word)
        # precomputes a set of the different characters
        sw = set(cw)
        # precomputes the "length" of the word vector
        lw = sqrt(sum(c*c for c in cw.values()))

        # return a tuple
        return cw, sw, lw

    def cosdis(v1, v2):
        # which characters are common to the two words?
        common = v1[1].intersection(v2[1])
        # by definition of cosine distance we have
        return sum(v1[0][ch]*v2[0][ch] for ch in common)/v1[2]/v2[2]
    
    list_of_keywords = ["Target", "Walmart","Best Buy","Apple","Kwik Trip",
        "Pick n Save","Speedway","McDonald's", "Panda Express",
        "Taco Bell","Potbelly","Burger King",
        "Zupas","MOD Pizza","Chipotle",
        "Noodles & Company","Dairy Queen","Subway","Visionworks",
        "Sunglass Hut","Starbucks","Jimmy Johns",
        "Chick-fil-A","Great Clips","Kohl's","Colder's"
        ,"KFC","Kopp's","Culver's","Leon's", 
        "Steinhafels","Home Depot", "OfficeMax",
        "Office Depot","Ashley","Zumiez","Yankee Candle",
        "Woops!","Windsor","Village Shoppe",
        "Victoria's Secret","Torrid","Tokyo Hut",
        "TJ-Maresults","Time Square","Tilly's","Things Remembered",
        "Telly's","Tanimar's Beauty",'T Mobile',
        "T Shirt City","Steak Escape",
        "Sports Images and More", "Sphinx Fashion", 
        "Spencer's", "Spectrum","Southridge Barber Shop",
        "Silver Choice","Shoe Dept. Encore","rue21","Round1",
        "Rogers & Hollands","Rocky Rococo","Rock America",
        "Rochelles Fashion Boutique","Robert Haack Jewelry Outlet","Renaisence","Sweet Basil",
        "JCPenney","The Children's Place","Menards","First Watch","Kohl's", "pieology", "Panera Bread",
        "Macy's","Lowes","California Pizza Kitchen","Kyoto","IKEA","Woodman's","Ghiradelli","Legoland","Papa Johns","Pick'n Save"]

    # Scan the Images
    imgd = Image.open(imng)
    #pytesseract.pytesseract.tesseract_cmd= 'tesseract/5.2.0/bin/tesseract'
    results = pytesseract.image_to_string(imgd, config="--psm 4")  
    # res var
    res = results.replace("/n", " " )
    res = res.replace("\n", " " )
    res = res.split(" ")
    res = list(res)
 # Find Total
    #First Way
    try:
        r = find_amounts(results)  
        r = max(r)
        wordIndex = results.index(str(r))
        nprice = r
    except:
        nprice = '' 
    #If 1st method doesn't work use this.    
    if nprice == '':
        n = ["Total","USD","Amount","Due","Total:","TOTAL","TOTAL:","Balance","BALANCE"]
        for x in n:
            try:
                wordIndex = res.index(str(n[x]))
                price = res[wordIndex + 1]
                nprice = re.findall(r'\d+\.\d+',price)
                nprice = ', '.join(nprice)
                if nprice != '':
                    break
            except:
                s = 0
        m = ["Total","Total:","TOTAL","TOTAL:","USD","Amount","Due","Balance","BALANCE"]
        for y in m:
            try:
                m = get_close_matches(y,res)
                m = str(m)[1:-1]
                m = m.replace("'", "")
                wordIndex1 = res.index(m)
                price = res[wordIndex1 + 1]
                nprice = re.findall(r'\d+\.\d+',price)
                nprice = ', '.join(nprice)
                if nprice != '':
                    break
            except:
                s = 0
        else:
            wordIndex = 1
            nprice = "nf"
            
            
 # Find Tip
    ftip = re.search("Tip", results)
    ftip1 = re.search("TIP", results)
    tipIndex = 0
    if wordIndex > tipIndex:
        nprice = nprice
    elif ftip:
        tipIndex = results.find("Tip")
        tip = results[tipIndex: tipIndex+10]
        ntip =  re.findall(r'\d+\.\d+',tip)
        ntip = ', '.join(ntip)
        ntotal = float(nprice) + float(ntip)
        nprice = ntotal
    elif ftip1:
        tipIndex = results.find("TIP")
        tip = results[tipIndex: tipIndex+10]
        ntip =  re.findall(r'\d+\.\d+',tip)
        ntip = ', '.join(ntip)
        ntotal = float(nprice) + float(ntip)
        nprice = ntotal
    else:
        print("No Tip")


 # Find Date
    d1 = re.findall(r'\d+/\d+/\d+',results)
    d2 = re.findall(r'\d+-\d+-\d+',results)
    d3 = re.findall(r'\d+-\D+-\d+',results)
    if d1:
        dte = d1
        dte = ', '.join(dte)

    elif d2:
        dte = d2
        dte = ', '.join(dte)

    elif d3:
        dte = d3
        dte = ', '.join(dte)
    else:
        dte = "nf"

 #Find Name
    uplist = [x.upper() for x in list_of_keywords] 
    lowlist =  [x.lower() for x in list_of_keywords]          

    if [ele for ele in list_of_keywords if(ele in results)]:
        name = [ele for ele in list_of_keywords if(ele in results)]
    elif [ele for ele in uplist if(ele in results)]:
        name = [ele for ele in uplist if(ele in results)]
    elif [ele for ele in lowlist if(ele in results)]:
        name = [ele for ele in lowlist if(ele in results)]
    else:
        name = [x for x in results.split() for y in list_of_keywords if cosdis(word2vec(x), word2vec(y)) > 0.9]
    
    if not name:
        name = results.partition('\n')[0]
        rr = [name,dte,nprice]
    else:
        namer = str(name)[1:-1]       
        rr= [namer,dte,nprice]
 # Main output
    df= pd.DataFrame([rr])
    df.columns = ['Name','Date','Total']
    # Uncomment after testing
    #if "nf" in rr:
        #df = pd.DataFrame(["Please rescan the image"])
        #df.columns = [f"Not working {img}"]  
    print(df) 
    with pd.ExcelWriter(saveto,mode="a",engine="openpyxl",if_sheet_exists="overlay") as writer:
        df.to_excel(writer, sheet_name="Sheet",header=None,startrow=writer.sheets["Sheet"].max_row,index=False)
    return df


# API Call
app = Flask(__name__)
@app.route('/find', methods=['GET'])
def fin():
    if request.method == 'GET':
        username = request.args['f']
        downlos = request.args['n']
        downlos = downlos.replace('"', '')
        username = username.replace('"', '')
        return {
            "RES" : find(username)
        }
    else:
        return {
            "RES" : "False"
        }
@app.route('/scan', methods=['GET'])
def scn():
    if request.method == 'GET':
        image = request.args['image']
        excel = request.args['excel']
        image = image[1:-1]
        excel = excel[1:-1]
        l = scan("8.jpg","n.xlsx")
        return {
            "RES" : f"True"
        }
    else:
        return {
            "RES" : "False"
        }
if __name__ == "__main__":
    # This is used when running locally only. When deploying to Google App
    # Engine, a webserver process such as Gunicorn will serve the app.
    app.run(host="127.0.0.1", port=8080, debug=True)
